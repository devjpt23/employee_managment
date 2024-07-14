import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

empDict = {}
readFile = pd.read_csv("C:\\Users\\25470\\Desktop\\MEETINGS\\meeting one fixing\\thisReal.csv")
redFill = PatternFill(patternType="solid", fgColor='C64747')

def creatingDict():
    for i in readFile._values:
        if i[1] in empDict:
            empDict[i[1]].append(i)
        else:
            empDict[i[1]] = []
            empDict[i[1]].append(i)            
creatingDict()

data = []
hoursPlacement = []

for key, value  in empDict.items():
    total_duration = sum([float(record[4]) for record in value if pd.notna(record[4])])
    for records in range(len(value)):
        data.append(value[records]) 

    sum_record = [None] * len(value[0])
    sum_record[4] = total_duration
    data.append(sum_record)
    hoursPlacement.append(len(data)+1)

columns = readFile.columns  
df = pd.DataFrame(data, columns=columns)

output_file = 'eachEmp.xlsx'
df.to_excel(output_file, index=False, engine='openpyxl')

workbook = load_workbook('eachEmp.xlsx')
sheet = workbook.active

for i in hoursPlacement:
    sheet[f'F{i}'] = (f"=E{i}/60")


endTime = sheet['D']
startTime = sheet['C']

startTimeArr = []
endTimeArr = []

def appendingIntoArr(arrName,columnName):
    for i in columnName[1:]:
        arrName.append(i.value)

appendingIntoArr(endTimeArr,endTime)
appendingIntoArr(startTimeArr,startTime)

print("comment : startTime : endTime")
def is_space(start,end):
    if (start and end) is None:
        pass
        
def is_blank(start,end,cell):
    if (start is not None) and (end is None):
        sheet[f'D{cell}'].fill = redFill
    else:
        pass

for i in range(len(endTimeArr)-1):  
    cell = i + 2 
    is_space(startTimeArr[i],endTimeArr[i])
    is_blank(startTimeArr[i],endTimeArr[i], cell)

workbook.save("finalFile.xlsx")#this file contains the file changes like summation of shifttime along with the conversion of time to hours and finally it contains coloring of cells with no shiftime entered
print('changes have been made!!')    