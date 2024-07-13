import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

empDict = {}
readFile = pd.read_csv("C:\\Users\\25470\\Desktop\\MEETINGS\\meeting one fixing\\thisReal.csv")
redFill = PatternFill(patternType="solid", fgColor='C64747')

for i in readFile._values:
    if i[1] in empDict:
        empDict[i[1]].append(i)
    else:
        empDict[i[1]] = []
        empDict[i[1]].append(i)

data = []
hoursPlacement = []
for key, value  in empDict.items():

    total_duration = sum([float(record[4]) for record in value if pd.notna(record[4])])

    for records in range(len(value)):
        data.append(value[records])        

    sum_record = [None] * len(value[0])
    sum_record[4] = total_duration
    data.append(sum_record)
    hoursPlacement.append(len(data)+1)


columns = readFile.columns  
df = pd.DataFrame(data, columns=columns)

output_file = 'eachEmp.xlsx'
df.to_excel(output_file, index=False, engine='openpyxl')

workbook = load_workbook('eachEmp.xlsx')
sheet = workbook.active

for i in hoursPlacement:
    sheet[f'F{i}'] = (f"=E{i}/60")

columnD = sheet['D']
for i in columnD:
    if i.value is None:
        i.fill = redFill

workbook.save("empHours.xlsx")
print('changes have been made!!')    