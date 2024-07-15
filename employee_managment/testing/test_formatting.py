import unittest
import pandas as pd
from formatting import creatingDict, is_blank
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class Testing(unittest.TestCase):

    def setUp(self):
        self.empDict = {}
        self.demoData = pd.DataFrame({
            'EmpID': [1, 1, 1, 2],
            'EmpName': ['Kevin Hart', 'Kevin Hart', 'Kevin Hart', 'Dwayne Johnson'],
            'StartTime': ['07:00', '07:00', '07:00', '07:00'],
            'EndTime': ['18:00', None, '19:00', '20:00'],
            'ShiftTime': ['660', 'N/A', '670', '680']
        })
        creatingDict(self.demoData, self.empDict)
        
        # Load the workbook and select the active sheet
        self.workbook = load_workbook('finalFile.xlsx')
        self.sheet = self.workbook.active
        
        # Define the red fill pattern
        self.redFill = PatternFill(patternType="solid", fgColor='C64747')

    def test_creatingDict(self):
        self.assertEqual(len(self.empDict['Kevin Hart']), 3)
        self.assertEqual(len(self.empDict['Dwayne Johnson']), 1)

    def test_is_blank(self):
        
        self.assertTrue(is_blank('07:00', None, 1))
        self.assertFalse(is_blank(None, None, 1))  
        self.assertFalse(is_blank('07:00', '18:00', 1))
        self.assertFalse(is_blank('07:00', '19:00', 1))

if __name__ == '__main__':
    unittest.main()
